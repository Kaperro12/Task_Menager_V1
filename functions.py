import openpyxl
import csv
from openpyxl.styles import (
    PatternFill, Font, Border, Side, Alignment
)
from tkinter import messagebox
from datetime import datetime
from tasks import TodoTask, InProgressTask, DoneTask

NAZWA_PLIKU = "Tasks.xlsx"
KOLORY = {
    "Do zrobienia": "#F28B82",
    "W trakcie": "#FBBC04",
    "Wykonane": "#81C995"
}


def sprawdz_lub_utworz_plik():
    try:
        return openpyxl.load_workbook(NAZWA_PLIKU)
    except FileNotFoundError:
        skoroszyt = openpyxl.Workbook()
        arkusz = skoroszyt.active
        arkusz.title = "Tasks"

        # Nagłówki kolumn
        headers = ["Title", "Status", "CreationDate", "Deadline"]
        arkusz.append(headers)

        # Formatowanie nagłówków
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cell in arkusz[1]:
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Szerokości kolumn
        arkusz.column_dimensions['A'].width = 40
        arkusz.column_dimensions['B'].width = 15
        arkusz.column_dimensions['C'].width = 20
        arkusz.column_dimensions['D'].width = 15

        skoroszyt.save(NAZWA_PLIKU)
        return skoroszyt


def zaladuj_zadania():
    skoroszyt = sprawdz_lub_utworz_plik()
    arkusz = skoroszyt.active
    zadania = []

    for wiersz in arkusz.iter_rows(min_row=2, values_only=True):
        if not wiersz or not wiersz[0]:  # Pomijaj puste wiersze
            continue

        tytul = wiersz[0]
        status = wiersz[1] if len(wiersz) > 1 else "Do zrobienia"

        try:
            data_utworzenia = (
                datetime.strptime(wiersz[2], "%Y-%m-%d %H:%M:%S")
                if len(wiersz) > 2 and wiersz[2]
                else datetime.now()
            )
        except ValueError:
            data_utworzenia = datetime.now()

        deadline = wiersz[3] if len(wiersz) > 3 and wiersz[3] else None

        if status == "Do zrobienia":
            zadanie = TodoTask(tytul)
        elif status == "W trakcie":
            zadanie = InProgressTask(tytul)
        elif status == "Wykonane":
            zadanie = DoneTask(tytul)
        else:
            zadanie = TodoTask(tytul)

        zadanie.data_utworzenia = data_utworzenia
        zadanie.deadline = deadline
        zadania.append(zadanie)

    return zadania


def zapisz_zadania(zadania):
    skoroszyt = openpyxl.Workbook()
    arkusz = skoroszyt.active
    arkusz.title = "Tasks"

    # Nagłówki
    headers = ["Title", "Status", "CreationDate", "Deadline"]
    arkusz.append(headers)

    # Formatowanie nagłówków
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in arkusz[1]:
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Szerokości kolumn
    arkusz.column_dimensions['A'].width = 40
    arkusz.column_dimensions['B'].width = 15
    arkusz.column_dimensions['C'].width = 20
    arkusz.column_dimensions['D'].width = 15

    # Dane zadań
    for zadanie in zadania:
        arkusz.append([
            zadanie.tytul,
            zadanie.status,
            zadanie.data_utworzenia.strftime("%Y-%m-%d %H:%M:%S"),
            zadanie.deadline if zadanie.deadline else None
        ])

        # Formatowanie komórek
        wypelnienie = PatternFill(
            start_color=KOLORY[zadanie.status][1:],
            end_color=KOLORY[zadanie.status][1:],
            fill_type="solid"
        )

        last_row = arkusz.max_row
        for row in arkusz.iter_rows(min_row=last_row, max_row=last_row):
            for cell in row:
                cell.fill = wypelnienie
                cell.border = thin_border

                if zadanie.status == "W trakcie":
                    cell.font = Font(bold=True)

                if cell.column_letter == 'D' and zadanie.deadline:
                    try:
                        deadline_date = datetime.strptime(zadanie.deadline, "%Y-%m-%d").date()
                        today = datetime.now().date()
                        if deadline_date < today and zadanie.status != "Wykonane":
                            cell.font = Font(color="FF0000", bold=True)
                    except ValueError:
                        pass

    skoroszyt.save(NAZWA_PLIKU)


def eksportuj_do_csv():
    zadania = zaladuj_zadania()
    with open('tasks_export.csv', mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file, delimiter=';')
        writer.writerow([
            'Tytuł', 'Status', 'Data utworzenia',
            'Termin', 'Status terminu'
        ])

        today = datetime.now().date()

        for zadanie in zadania:
            deadline_status = ""
            if zadanie.deadline:
                try:
                    deadline_date = datetime.strptime(zadanie.deadline, "%Y-%m-%d").date()
                    if zadanie.status == "Wykonane":
                        deadline_status = "Wykonane przed terminem"
                    elif deadline_date < today:
                        deadline_status = "Zaległe"
                    else:
                        deadline_status = "W terminie"
                except ValueError:
                    deadline_status = "Nieprawidłowy format daty"

            writer.writerow([
                zadanie.tytul,
                zadanie.status,
                zadanie.data_utworzenia.strftime("%Y-%m-%d %H:%M"),
                zadanie.deadline if zadanie.deadline else "Brak",
                deadline_status
            ])


def generuj_statystyki():
    zadania = zaladuj_zadania()
    today = datetime.now().date()

    stats = {
        'total': len(zadania),
        'todo': 0,
        'in_progress': 0,
        'done': 0,
        'overdue': 0
    }

    for zadanie in zadania:
        if zadanie.status == "Do zrobienia":
            stats['todo'] += 1
        elif zadanie.status == "W trakcie":
            stats['in_progress'] += 1
        elif zadanie.status == "Wykonane":
            stats['done'] += 1

        if zadanie.deadline and zadanie.status != "Wykonane":
            try:
                deadline_date = datetime.strptime(zadanie.deadline, "%Y-%m-%d").date()
                if deadline_date < today:
                    stats['overdue'] += 1
            except ValueError:
                pass

    return stats


def dodaj_zadanie(tytul, deadline=None, app_instance=None):
    if not tytul or len(tytul.strip()) == 0:
        messagebox.showwarning("Ostrzeżenie", "Tytuł zadania nie może być pusty")
        return False

    if len(tytul) > 100:
        messagebox.showwarning(
            "Ostrzeżenie",
            "Tytuł zadania jest zbyt długi (max 100 znaków)"
        )
        return False

    zadania = zaladuj_zadania()

    # Sprawdź duplikaty (case-insensitive)
    if any(z.tytul.lower() == tytul.lower() for z in zadania):
        messagebox.showwarning(
            "Ostrzeżenie",
            "Zadanie o podanym tytule już istnieje"
        )
        return False

    # Utwórz nowe zadanie
    nowe_zadanie = TodoTask(tytul)
    if deadline:
        nowe_zadanie.deadline = deadline

    zadania.append(nowe_zadanie)
    zapisz_zadania(zadania)

    if app_instance:
        app_instance.odswiez_liste_zadan()

    return True


def edytuj_zadanie(stary_tytul, nowy_tytul, app_instance=None):
    if not nowy_tytul or len(nowy_tytul.strip()) == 0:
        messagebox.showwarning("Ostrzeżenie", "Tytuł zadania nie może być pusty")
        return False

    if len(nowy_tytul) > 100:
        messagebox.showwarning(
            "Ostrzeżenie",
            "Tytuł zadania jest zbyt długi (max 100 znaków)"
        )
        return False

    zadania = zaladuj_zadania()

    # Znajdź zadanie do edycji
    zadanie_do_edycji = None
    for z in zadania:
        if z.tytul == stary_tytul:
            zadanie_do_edycji = z
            break

    if not zadanie_do_edycji:
        messagebox.showwarning(
            "Ostrzeżenie",
            "Nie znaleziono zadania do edycji"
        )
        return False

    # Sprawdź czy nowy tytuł nie koliduje z innym zadaniem
    if (nowy_tytul.lower() != stary_tytul.lower() and
            any(z.tytul.lower() == nowy_tytul.lower() for z in zadania)):
        messagebox.showwarning(
            "Ostrzeżenie",
            "Zadanie o podanym tytule już istnieje"
        )
        return False

    # Zaktualizuj zadanie
    zadanie_do_edycji.tytul = nowy_tytul
    zapisz_zadania(zadania)

    if app_instance:
        app_instance.odswiez_liste_zadan()

    return True


def rozpocznij_zadanie(tytul, app_instance=None):
    zadania = zaladuj_zadania()
    for zadanie in zadania:
        if zadanie.tytul == tytul and zadanie.status == "Do zrobienia":
            zadanie.status = "W trakcie"
            zapisz_zadania(zadania)

            if app_instance:
                app_instance.odswiez_liste_zadan()

            return True

    messagebox.showwarning(
        "Ostrzeżenie",
        "Nie można rozpocząć tego zadania (nie istnieje lub ma niewłaściwy status)"
    )
    return False


def zakoncz_zadanie(tytul, app_instance=None):
    zadania = zaladuj_zadania()
    for zadanie in zadania:
        if zadanie.tytul == tytul and zadanie.status == "W trakcie":
            zadanie.status = "Wykonane"
            zapisz_zadania(zadania)

            if app_instance:
                app_instance.odswiez_liste_zadan()

            return True

    messagebox.showwarning(
        "Ostrzeżenie",
        "Nie można zakończyć tego zadania (nie istnieje lub ma niewłaściwy status)"
    )
    return False


def usun_zadanie(tytul, app_instance=None):
    zadania = zaladuj_zadania()
    nowe_zadania = [z for z in zadania if z.tytul != tytul]

    if len(nowe_zadania) < len(zadania):
        zapisz_zadania(nowe_zadania)

        if app_instance:
            app_instance.odswiez_liste_zadan()

        return True

    messagebox.showwarning(
        "Ostrzeżenie",
        "Nie znaleziono zadania do usunięcia"
    )
    return False